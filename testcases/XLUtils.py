import openpyxl as op


def get_row_count(file, sheetname):
    workbook = op.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.max_row


def get_col_count(file, sheetname):
    workbook = op.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.max_column


def read_data(file, sheetname, rownum, columnnum):
    workbook = op.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnnum).value


def write_data(file, sheetname, rownum, columnnum, data):
    workbook = op.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnnum).value = data
    workbook.save(file)






